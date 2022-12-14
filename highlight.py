import fitz
import openpyxl
import json

unfound = {}

def hightligh_in_page(doc, index, keyword):
    found = False
    for i in range(50):
        page = doc[min(len(doc)-1, index-25+i)]

        text_instances = page.search_for(keyword)
        for inst in text_instances:
            found = True
            print("found istance on page:", index-25+i)
            highlight = page.add_highlight_annot(inst)
            highlight.update()
    if not found:
        if keyword in unfound:
            unfound[keyword].append(index)
        else:
            num = []
            num.append(index)
            unfound[keyword] = num

### READ IN PDF
doc = fitz.open("history-of-western-phil.pdf")

# for page in doc:
#     ### SEARCH
#     text = "information"
#     text_instances = page.searchFor(text)

#     ### HIGHLIGHT
#     for inst in text_instances:
#         highlight = page.addHighlightAnnot(inst)
#         highlight.update()
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("index.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in dataframe1.iter_rows(1, dataframe1.max_row):
    keyword = row[1].value
    print("checking for keyword:", keyword)
    for col in range(3, dataframe1.max_column):
        index = row[col].value
        if isinstance(index, int):
            print("page is:", row[col].value)
            hightligh_in_page(doc, index, keyword)

### OUTPUT
doc.save("highlighted.pdf", garbage=4, deflate=True, clean=True)
with open('unfound.txt', 'w') as convert_file:
     convert_file.write(json.dumps(unfound))

